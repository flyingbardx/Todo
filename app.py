# Import packages.
from tradingview_ta import TA_Handler, Interval, Exchange,get_multiple_analysis
import time
from openpyxl import load_workbook
wb = load_workbook("TRADE.xlsx")
ws = wb.active
# Store the last order.
start_row = 2
symbol = 'BTCUSDT'
screener = "crypto"
exchange = "BINANCE"

last_order = "sell"

# Instantiate TA_Handler.


# Repeat forever.
while True:
    # Retrieve recommendation.
    print('Fetching data')

    rec = TA_Handler(
	    symbol=symbol,
	    exchange=exchange,
	    screener=screener,
	    interval=Interval.INTERVAL_1_MINUTE,
	).get_analysis().summary["RECOMMENDATION"]

    # rec = handler.get_analysis().summary
    analysis = get_multiple_analysis(screener=screener, interval=Interval.INTERVAL_1_MINUTE, symbols=[exchange+":BTCUSDT"])
    indicators = (analysis['BINANCE:BTCUSDT'].indicators)


    # Create a buy order if the recommendation is "BUY" or "STRONG_BUY" and the last order is "sell".
    # Create a sell order if the recommendation is "SELL" or "STRONG_SELL" and the last order is "buy".
    
    print("CURRENT STATUS "+rec)
    if "BUY" in rec and last_order == "sell":
    	ws.cell(row=start_row, column= 1 ).value = start_row-1
    	ws.cell(row=start_row, column= 2 ).value = symbol
    	ws.cell(row=start_row, column= 4 ).value = indicators['close']
    	ws.cell(row=start_row, column= 3 ).value = "BUY"
    	last_order = "buy"
    	start_row += 1
    elif "SELL" in rec and last_order == "buy":
    	ws.cell(row=start_row, column= 1 ).value = start_row-1
    	ws.cell(row=start_row, column= 2 ).value = symbol
    	ws.cell(row=start_row, column= 4 ).value = indicators['close']
    	ws.cell(row=start_row, column= 3 ).value = "SELL"
    	last_order = "sell"
    	start_row += 1
    	
    wb.save("TRADE.xlsx")
    time.sleep(60)